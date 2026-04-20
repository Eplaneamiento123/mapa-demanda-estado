"""
Análisis de Demanda - Fibra Óptica B2B (Versión Completa y Original)
Ejecución Local - Requiere conexión VPN activa
"""

import os
import pyodbc
import pandas as pd
import numpy as np
import geopandas as gpd
import folium
from folium.plugins import HeatMap, MarkerCluster
import requests, math, warnings, hashlib, json
from scipy.sparse.csgraph import minimum_spanning_tree
from scipy.spatial.distance import cdist
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ==========================================
# 0. DIRECTORIO DE SALIDA
# ==========================================
# NOTA: Si deseas cambiar la carpeta de salida, modifica OUTPUT_DIR aquí abajo.
# El mapa y el Excel se guardan en la misma carpeta donde está este script.
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
os.makedirs(OUTPUT_DIR, exist_ok=True)
SCRIPT_DIR = OUTPUT_DIR

# ==========================================
# 1. CONFIGURACIÓN DE RUTAS LOCALES
# ==========================================
RUTA_CLIENTES_ESTADO = r"C:\Users\csegil\OneDrive - On Empresas S.A.C\Carlos Andres\09. Demanda Potencial\02. ANALISIS CLIENTES\00. PROYECTO ESTADO\INFORMACION A ANALIZAR\CLIENTES ESTADO.xlsx"
RUTA_NODOS           = r"C:\Users\csegil\OneDrive - On Empresas S.A.C\Carlos Andres\09. Demanda Potencial\02. ANALISIS CLIENTES\00. PROYECTO ESTADO\INFORMACION A ANALIZAR\Nodos y responsables_v2 1.xlsx"

# ==========================================
# 2. PARÁMETROS DEL NEGOCIO
# ==========================================
BUFFER_METROS = 750
NODOS_BUFFER_M = 1000  

COL_ESTADO = {
    'item': 'ITEM', 'direccion': 'DIRECCION', 'ruc': 'RUC', 'cliente': 'CLIENTE',
    'sede': 'SEDE', 'distrito': 'DISTRITO', 'provincia': 'PROVINCIA', 
    'depto': 'DEPARTAMENTO', 'latitud': 'LATITUD', 'longitud': 'LONGITUD'
}
COL_DEMANDA = {
    'item': 'ITEM', 'direccion': 'origen_direccion', 'ruc': 'ruc', 'cliente': 'cliente',
    'fuente': 'FUENTE', 'territorio': 'territorio', 'distrito': 'DISTRITO',
    'provincia': 'PROVINCIA', 'latitud': 'LATITUD', 'longitud': 'LONGITUD'
}

PALETA = ['#E63946','#2196F3','#4CAF50','#FF9800','#9C27B0','#00BCD4','#F44336','#795548','#607D8B','#E91E63']
COLOR_ENTIDAD = {'ANA': '#6D4C41', 'CONECTAMEF': '#2196F3', 'ONP': '#4CAF50', 'ONPE': '#FF9800', 'ORS': '#9C27B0', 'PNP': '#00BCD4', 'SUNAT': '#E53935', 'SENCICO': '#757575'}
HEX_FUENTE = {'SUNAT': '#1565C0', 'PREFACTIBILIDAD': '#2E7D32', 'COMPETENCIA': '#E65100', 'COMPETIDORES': '#E65100', 'PREFA': '#2E7D32'}

FACTOR_REGION = {
    'TUMBES': 1.35, 'PIURA': 1.35, 'LAMBAYEQUE': 1.35, 'LA LIBERTAD': 1.35, 'ANCASH': 1.45, 'LIMA': 1.35,
    'ICA': 1.35, 'AREQUIPA': 1.40, 'MOQUEGUA': 1.40, 'TACNA': 1.38, 'CALLAO': 1.35,
    'CAJAMARCA': 1.55, 'AMAZONAS': 1.50, 'HUANUCO': 1.50, 'PASCO': 1.55, 'JUNIN': 1.50, 'HUANCAVELICA': 1.55,
    'AYACUCHO': 1.55, 'APURIMAC': 1.55, 'CUSCO': 1.50, 'PUNO': 1.50,
    'LORETO': 1.45, 'UCAYALI': 1.45, 'MADRE DE DIOS': 1.45, 'SAN MARTIN': 1.45,
}
FACTOR_CALLE_DEFAULT = 1.40

def get_factor(depto):
    if not depto or str(depto).strip() in ('', 'nan', 'None'): return FACTOR_CALLE_DEFAULT
    return FACTOR_REGION.get(str(depto).upper().strip(), FACTOR_CALLE_DEFAULT)

# ==========================================
# 3. FUNCIONES ESPACIALES Y DESCARGA GEOJSON
# ==========================================
def haversine_matrix(coords):
    lat = coords[:, 0]; lon = coords[:, 1]
    dlat = lat[:, None] - lat[None, :]
    dlon = lon[:, None] - lon[None, :]
    a = (np.sin(dlat/2)**2 + np.cos(lat[:, None]) * np.cos(lat[None, :]) * np.sin(dlon/2)**2)
    return 6371000 * 2 * np.arctan2(np.sqrt(a), np.sqrt(1-a))

def calcular_mst_km(sede_lat, sede_lon, empresas_lat, empresas_lon):
    if len(empresas_lat) == 0: return 0.0
    lats = np.radians(np.array([sede_lat] + list(empresas_lat)))
    lons = np.radians(np.array([sede_lon] + list(empresas_lon)))
    coords = np.column_stack([lats, lons])
    dist_matrix = haversine_matrix(coords) * FACTOR_CALLE_DEFAULT
    mst = minimum_spanning_tree(dist_matrix)
    return float(mst.sum())

def hav_simple(lat1, lon1, lat2, lon2):
    R = 6371000
    la1 = np.radians(lat1); lo1 = np.radians(lon1)
    la2 = np.radians(lat2); lo2 = np.radians(lon2)
    a = (np.sin((la2-la1)/2)**2 + np.cos(la1)*np.cos(la2)*np.sin((lo2-lo1)/2)**2)
    return R * 2 * np.arcsin(np.sqrt(np.clip(a, 0, 1)))

print('🗺️ Descargando límites geográficos del Perú...')
URLS_DISTRITOS = ['https://raw.githubusercontent.com/rionlabs/peru-geojson/main/peru_distritos.geojson']
URLS_PROVINCIAS = ['https://raw.githubusercontent.com/rionlabs/peru-geojson/main/peru_provincias.geojson']

def intentar_descarga(urls, nombre):
    for url in urls:
        try:
            r = requests.get(url, timeout=10)
            if r.status_code == 200:
                data = r.json()
                print(f'   ✅ {nombre}: cargado correctamente')
                return data
        except: pass
    return None

GEOJSON_DISTRITOS  = intentar_descarga(URLS_DISTRITOS, 'Distritos')
GEOJSON_PROVINCIAS = intentar_descarga(URLS_PROVINCIAS, 'Provincias')

def detectar_campo(geojson, candidatos):
    if not geojson: return None
    props = geojson['features'][0].get('properties', {})
    for c in candidatos:
        if c in props: return c
    return list(props.keys())[0] if props else None

CAMPO_DISTRITO  = detectar_campo(GEOJSON_DISTRITOS, ['NOMBDIST','nomDist','nombre_dis','distrito','DISTRITO','name','NAME'])
CAMPO_PROVINCIA = detectar_campo(GEOJSON_PROVINCIAS, ['NOMBPROV','nomProv','nombre_pro','provincia','PROVINCIA','name','NAME'])
COLORES_DIST = ['#E3F2FD','#E8F5E9','#FFF3E0','#F3E5F5','#E0F7FA','#FBE9E7','#F1F8E9','#EDE7F6','#E8EAF6','#FCE4EC']

# ==========================================
# 4. EXTRACCIÓN Y LIMPIEZA DE DATOS
# ==========================================
print("\n📡 Iniciando Análisis de Demanda B2B...")
try:
    conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=10.1.3.85;DATABASE=win_empresas;Trusted_Connection=yes;')
    df_demanda_raw = pd.read_sql("""
        SELECT CODIGO AS ITEM, CASE FUENTE WHEN 'PREFA' THEN 'PREFACTIBILIDAD' ELSE FUENTE END AS FUENTE,
        RUC AS ruc, RAZON_SOCIAL AS cliente, DIRECCION_COMPLETA AS origen_direccion, TERITORIO_CORREGIDO AS territorio, 
        DEPARTAMENTO, PROVINCIA, DISTRITO, LATITUD, LONGITUD, TIPO_SERVICIO AS tipo_servicio, FACTURACION
        FROM Demanda.vw_Demanda_Proyecto_Estado
    """, conn)
    conn.close()
    print(f"✅ Demanda B2B SQL cargada: {len(df_demanda_raw):,} registros")
except Exception as e:
    print(f"❌ Error conectando a SQL: {e}"); exit()

try:
    df_estado_raw = pd.read_excel(RUTA_CLIENTES_ESTADO)
    print(f"✅ Sedes Excel cargadas: {len(df_estado_raw)} filas")
except Exception as e:
    print(f"❌ Error leyendo Excel de sedes: {e}"); exit()

def limpiar_coordenadas(df, col_lat, col_lon):
    mask_inv = (df[col_lat].isna() | df[col_lon].isna() | (df[col_lat] < -18.5) | (df[col_lat] > -0.5) | (df[col_lon] < -82.0) | (df[col_lon] > -68.0))
    return df[~mask_inv].reset_index(drop=True)

df_estado = limpiar_coordenadas(df_estado_raw, COL_ESTADO['latitud'], COL_ESTADO['longitud'])
df_demanda = limpiar_coordenadas(df_demanda_raw, COL_DEMANDA['latitud'], COL_DEMANDA['longitud'])

# ==========================================
# 5. GEOPROCESAMIENTO Y CRUCE (SPATIAL JOIN)
# ==========================================
print("\n🗺️ Procesando buffers y cruce espacial...")
lat_e = COL_ESTADO['latitud']; lon_e = COL_ESTADO['longitud']
lat_d = COL_DEMANDA['latitud']; lon_d = COL_DEMANDA['longitud']

df_estado[COL_ESTADO['item']] = df_estado[COL_ESTADO['item']].astype(str).str.strip()
df_demanda[COL_DEMANDA['item']] = df_demanda[COL_DEMANDA['item']].astype(str).str.strip()

gdf_sedes = gpd.GeoDataFrame(df_estado, geometry=gpd.points_from_xy(df_estado[lon_e], df_estado[lat_e]), crs='EPSG:4326').to_crs('EPSG:32718')
gdf_dem = gpd.GeoDataFrame(df_demanda, geometry=gpd.points_from_xy(df_demanda[lon_d], df_demanda[lat_d]), crs='EPSG:4326').to_crs('EPSG:32718')

def limpiar_sede(row):
    n = str(row[COL_ESTADO['sede']]).strip()
    return row[COL_ESTADO['direccion']][:40] if n in ('','-','nan','None') else n

gdf_sedes['_NOMBRE_SEDE'] = gdf_sedes.apply(limpiar_sede, axis=1)
gdf_buffers = gdf_sedes.copy()
gdf_buffers['geometry'] = gdf_sedes.geometry.buffer(BUFFER_METROS)

cols_sede = {COL_ESTADO['item']: '_SEDE_ITEM', COL_ESTADO['cliente']: '_SEDE_ENTIDAD', COL_ESTADO['direccion']: '_SEDE_DIR', lat_e: '_SEDE_LAT', lon_e: '_SEDE_LON', '_NOMBRE_SEDE': '_SEDE_NOMBRE'}
gdf_buffers_join = gdf_buffers[['geometry'] + list(cols_sede.keys())].rename(columns=cols_sede)

joined = gpd.sjoin(gdf_dem.reset_index().rename(columns={'index': '_idx_dem'}), gdf_buffers_join, how='left', predicate='within')

la1 = np.radians(joined['_SEDE_LAT'].fillna(0).values); lo1 = np.radians(joined['_SEDE_LON'].fillna(0).values)
la2 = np.radians(joined[lat_d].values); lo2 = np.radians(joined[lon_d].values)
dlat = la2 - la1; dlon = lo2 - lo1
a = np.sin(dlat/2)**2 + np.cos(la1)*np.cos(la2)*np.sin(dlon/2)**2

_factores = joined['_SEDE_ITEM'].map(df_estado.set_index(COL_ESTADO['item'])[COL_ESTADO['depto']].apply(get_factor)).fillna(FACTOR_CALLE_DEFAULT)
joined['_DIST_M'] = np.where(joined['_SEDE_ITEM'].notna(), 6371000 * 2 * np.arctan2(np.sqrt(a), np.sqrt(1-a)) * _factores, np.inf)

joined_dedup = joined.sort_values('_DIST_M').drop_duplicates(subset='_idx_dem', keep='first').copy()
joined_dedup['EN_BUFFER'] = np.where(joined_dedup['_SEDE_ITEM'].notna(), 'DENTRO', 'FUERA')
joined_dedup['ENTIDAD_ESTADO'] = joined_dedup['_SEDE_ENTIDAD'].fillna('')
joined_dedup['SEDE_ITEM'] = joined_dedup['_SEDE_ITEM'].fillna('').astype(str).str.strip()
joined_dedup['SEDE_NOMBRE'] = joined_dedup['_SEDE_NOMBRE'].fillna('')
joined_dedup['SEDE_DIRECCION'] = joined_dedup['_SEDE_DIR'].fillna('')
joined_dedup['DISTANCIA_METROS'] = joined_dedup['_DIST_M'].replace([np.inf, -np.inf], np.nan).fillna(0).round().astype(int)

df_demanda_full = joined_dedup.drop(columns=['geometry','_idx_dem','_DIST_M','index_right','_SEDE_LAT','_SEDE_LON','_SEDE_DIR','_SEDE_ENTIDAD','_SEDE_NOMBRE'], errors='ignore').copy()
df_det = df_demanda_full[df_demanda_full['EN_BUFFER'] == 'DENTRO'].copy()

# --- CAPA VISUAL SIN DEDUPLICACIÓN + JITTER ---
# df_mapa_visual incluye TODAS las relaciones empresa-sede (sin dedup por empresa),
# para que al filtrar/ocultar una entidad los puntos compartidos sigan visibles
# en las otras entidades que los contienen.
# El jitter desplaza puntos solapados ~10m aleatoriamente para que sean distinguibles.
_rng = np.random.default_rng(42)
_joined_visual = joined[joined['_SEDE_ITEM'].notna()].copy()
_joined_visual['DISTANCIA_METROS'] = _joined_visual['_DIST_M'].replace([np.inf,-np.inf], np.nan).fillna(0).round().astype(int)
_joined_visual['ENTIDAD_ESTADO']   = _joined_visual['_SEDE_ENTIDAD'].fillna('')
_joined_visual['SEDE_ITEM']        = _joined_visual['_SEDE_ITEM'].fillna('').astype(str).str.strip()
_joined_visual['SEDE_NOMBRE']      = _joined_visual['_SEDE_NOMBRE'].fillna('')
_joined_visual['SEDE_DIRECCION']   = _joined_visual['_SEDE_DIR'].fillna('')
df_mapa_visual = _joined_visual.drop(columns=['geometry','_idx_dem','_DIST_M','index_right','_SEDE_LAT','_SEDE_LON','_SEDE_DIR','_SEDE_ENTIDAD','_SEDE_NOMBRE'], errors='ignore').copy()

# Jitter por grupo (SEDE_ITEM + coordenada): si N empresas comparten ubicación
# exacta, se mantiene la primera en su sitio y las demás se distribuyen en un
# pequeño círculo ~15m alrededor para que todas sean visibles.
df_mapa_visual = df_mapa_visual.reset_index(drop=True)
_grp_cols = ['SEDE_ITEM', lat_d, lon_d]
_df_tmp = df_mapa_visual[_grp_cols].copy()
_df_tmp[lat_d] = _df_tmp[lat_d].round(6)
_df_tmp[lon_d] = _df_tmp[lon_d].round(6)
_df_tmp['_rank'] = _df_tmp.groupby(_grp_cols).cumcount()
# ~0.00013 grados ≈ 15 metros
_jitter_r = 0.00013
_mask = _df_tmp['_rank'] > 0
if _mask.sum() > 0:
    # Desplazamiento en círculo: ángulo basado en rank, radio constante
    _theta = _df_tmp.loc[_mask, '_rank'].values * (2 * np.pi / 6)  # 6 posiciones alrededor
    df_mapa_visual.loc[_mask, lat_d] = df_mapa_visual.loc[_mask, lat_d].values + _jitter_r * np.sin(_theta)
    df_mapa_visual.loc[_mask, lon_d] = df_mapa_visual.loc[_mask, lon_d].values + _jitter_r * np.cos(_theta)

entidades = sorted(df_estado[COL_ESTADO['cliente']].unique())
color_map  = {e: COLOR_ENTIDAD.get(e, PALETA[i % len(PALETA)]) for i, e in enumerate(entidades)}
res_rows = []

for _, sede in df_estado.iterrows():
    ent = sede[COL_ESTADO['cliente']]
    nombre_sede = limpiar_sede(sede)
    item_sede_str = str(sede[COL_ESTADO['item']]).strip()
    sub = df_det[df_det['SEDE_ITEM'] == item_sede_str]

    mst_km = 0.0
    factor_sede = get_factor(sede.get(COL_ESTADO['depto'], ''))
    if len(sub) > 0:
        mst_km = round((calcular_mst_km(sede[lat_e], sede[lon_e], sub[lat_d].values, sub[lon_d].values) * factor_sede) / 1000, 3)

    conteo = sub[COL_DEMANDA['fuente']].value_counts().to_dict() if len(sub) > 0 else {}
    res_rows.append({
        'ENTIDAD': ent, 'SEDE_ITEM': item_sede_str, 'SEDE_NOMBRE': nombre_sede,
        'SEDE_DIRECCION': sede[COL_ESTADO['direccion']], 'SEDE_DISTRITO': sede.get(COL_ESTADO['distrito'], ''),
        'SEDE_PROVINCIA': sede.get(COL_ESTADO['provincia'], ''), 'SEDE_DEPTO': sede.get(COL_ESTADO['depto'], ''),
        'SEDE_LAT': sede[lat_e], 'SEDE_LON': sede[lon_e], 'TOTAL_EMPRESAS': len(sub), 
        **{f'FUENTE_{k}': v for k, v in conteo.items()}, 'METRAJE_MST_KM': mst_km
    })

df_res = pd.DataFrame(res_rows)
df_ent = df_res.groupby('ENTIDAD').agg(N_SEDES=('SEDE_ITEM','count'), TOTAL_EMPRESAS=('TOTAL_EMPRESAS','sum'), METRAJE_MST_KM=('METRAJE_MST_KM','sum')).reset_index()

# ==========================================
# 6. ANÁLISIS DE NODOS (Opcional)
# ==========================================
nodos_analisis = []
df_demanda_nodos = None

if RUTA_NODOS:
    print("\n🏢 Procesando Nodos de Fibra Óptica...")
    try:
        _df_n = pd.read_excel(RUTA_NODOS, sheet_name='Nodos', header=0)
        _df_n.columns = [str(c).strip().replace('\xa0','') for c in _df_n.columns]
        _df_n['LATITUD'] = pd.to_numeric(_df_n['LATITUD'], errors='coerce')
        _df_n['LONGITUD'] = pd.to_numeric(_df_n['LONGITUD'], errors='coerce')
        _df_n = _df_n.dropna(subset=['LATITUD','LONGITUD']).reset_index(drop=True)

        _lats_e, _lons_e = df_estado[lat_e].values, df_estado[lon_e].values
        _lats_d, _lons_d = df_demanda[lat_d].values, df_demanda[lon_d].values
        _n_dem = len(df_demanda)
        _n_nodos = len(_df_n)

        _mat = np.zeros((_n_dem, _n_nodos))
        for _j, (_, _nr) in enumerate(_df_n.iterrows()):
            _mat[:,_j] = hav_simple(float(_nr['LATITUD']), float(_nr['LONGITUD']), _lats_d, _lons_d)

        _min_dist = _mat.min(axis=1)
        _nodo_asig = _mat.argmin(axis=1)
        _en_algun_nodo = _min_dist <= NODOS_BUFFER_M

        _cubierta_por_sede = np.zeros(_n_dem, dtype=bool)
        for _ei in range(_n_dem):
            _d_sedes = hav_simple(_lats_d[_ei], _lons_d[_ei], _lats_e, _lons_e)
            _cubierta_por_sede[_ei] = (_d_sedes <= BUFFER_METROS).any()

        _det_rows = []
        for _ei in range(_n_dem):
            if not _en_algun_nodo[_ei]: continue
            _j = _nodo_asig[_ei]
            _nr = _df_n.iloc[_j]
            _row = df_demanda.iloc[_ei]
            _det_rows.append({
                'NODO': str(_nr.get('Nombre_Nodo', '-')), 'NODO_ESTADO': str(_nr.get('Nodo','-')),
                'NODO_STATUS': str(_nr.get('Status','-')), 'DIST_AL_NODO_M': round(float(_min_dist[_ei]),1),
                'CATEGORIA': 'YA CUBIERTA' if _cubierta_por_sede[_ei] else 'CLIENTE POTENCIAL',
                'CLIENTE': _row.get(COL_DEMANDA['cliente'],'-'), 'RUC': _row.get(COL_DEMANDA['ruc'],'-'),
                'DISTRITO': _row.get(COL_DEMANDA['distrito'],'-'), 'LATITUD': _lats_d[_ei], 'LONGITUD': _lons_d[_ei]
            })
        df_demanda_nodos = pd.DataFrame(_det_rows)

        for _j, (_, _nr) in enumerate(_df_n.iterrows()):
            _mask_nodo = (_nodo_asig == _j) & _en_algun_nodo
            _ds = hav_simple(float(_nr['LATITUD']), float(_nr['LONGITUD']), _lats_e, _lons_e)
            nodos_analisis.append({
                'nombre': str(_nr.get('Nombre_Nodo','-')), 'estado': str(_nr.get('Nodo','-')),
                'status': str(_nr.get('Status','-')), 'depto': str(_nr.get('Departamento','-')),
                'ciudad': str(_nr.get('Ciudad','-')), 'direccion': str(_nr.get('Direccion','-'))[:60],
                'lat': float(_nr['LATITUD']), 'lon': float(_nr['LONGITUD']),
                'sedes_5km': int((_ds <= NODOS_BUFFER_M).sum()), 'emp_5km': int(_mask_nodo.sum()),
                'emp_ya_cubiertas': int((_mask_nodo & _cubierta_por_sede).sum()), 'emp_potenciales': int((_mask_nodo & ~_cubierta_por_sede).sum())
            })
        print(f"✅ Nodos procesados: {len(nodos_analisis)}")
    except Exception as e:
        print(f"❌ Error procesando nodos: {e}")

# ==========================================
# 7. GENERACIÓN DE MAPA HTML (FOLIUM - VERSIÓN ORIGINAL COMPLETA)
# ==========================================
print("\n🌐 Generando mapa interactivo HTML...")
todos_lats = list(df_estado[lat_e]) + list(df_demanda[lat_d])
todos_lons = list(df_estado[lon_e]) + list(df_demanda[lon_d])
sw = [min(todos_lats)-0.5, min(todos_lons)-0.5]
ne = [max(todos_lats)+0.5, max(todos_lons)+0.5]

mapa = folium.Map(tiles=None)
mapa.get_root().html.add_child(folium.Element("<meta name='viewport' content='width=device-width, initial-scale=1.0, maximum-scale=1.0'>"))
# Fondos: limpio por defecto, satélite para ver edificios/zonas, oscuro para contraste
folium.TileLayer('CartoDB positron', name='🗺️ Limpio', show=True).add_to(mapa)
folium.TileLayer(
    tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
    name='🛰️ Satélite',
    attr='Esri World Imagery',
    show=False
).add_to(mapa)
folium.TileLayer(
    tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Street_Map/MapServer/tile/{z}/{y}/{x}',
    name='🏙️ Calles (tipo Google)',
    attr='Esri World Street Map',
    show=False
).add_to(mapa)
folium.TileLayer('OpenStreetMap', name='🛣️ OpenStreetMap', show=False).add_to(mapa)
folium.TileLayer('CartoDB dark_matter', name='🌙 Oscuro', show=False).add_to(mapa)
mapa.fit_bounds([sw, ne])

if GEOJSON_PROVINCIAS and CAMPO_PROVINCIA:
    g_prov = folium.FeatureGroup(name='Limites Provincias', show=False)
    folium.GeoJson(GEOJSON_PROVINCIAS,
        style_function=lambda x: {'fillColor':'#FFF8E1','color':'#5D4037','weight':2,'fillOpacity':0.15},
        highlight_function=lambda x: {'fillColor':'#FF8F00','color':'#E65100','weight':3,'fillOpacity':0.4},
        tooltip=folium.GeoJsonTooltip(fields=[CAMPO_PROVINCIA], aliases=['Provincia:'], style='font-family:Arial;font-size:12px;background:white;padding:5px;border-radius:5px', sticky=True)
    ).add_to(g_prov)
    g_prov.add_to(mapa)

if GEOJSON_DISTRITOS and CAMPO_DISTRITO:
    g_dist = folium.FeatureGroup(name='Distritos', show=False)
    def color_distrito(props):
        idx = int(hashlib.md5(str(props.get(CAMPO_DISTRITO,'')).encode()).hexdigest(),16) % len(COLORES_DIST)
        return COLORES_DIST[idx]
    folium.GeoJson(GEOJSON_DISTRITOS,
        style_function=lambda x: {'fillColor':color_distrito(x['properties']),'color':'#1565C0','weight':1,'fillOpacity':0.3},
        highlight_function=lambda x: {'fillColor':'#1565C0','color':'#0D47A1','weight':2.5,'fillOpacity':0.55},
        tooltip=folium.GeoJsonTooltip(fields=[CAMPO_DISTRITO], aliases=['Distrito:'], style='font-family:Arial;font-size:12px;background:white;padding:5px;border-radius:5px', sticky=True)
    ).add_to(g_dist)
    g_dist.add_to(mapa)

g_fue = folium.FeatureGroup(name='Demanda FUERA de buffer', show=False)
# Clustering: los puntos fuera del buffer se agrupan cuando están alejados.
# Al hacer zoom se despliegan. Evita saturación visual de miles de puntos.
_cluster_fue = MarkerCluster(
    name='_cluster_fuera',
    options={'maxClusterRadius': 40, 'showCoverageOnHover': False, 'spiderfyOnMaxZoom': True}
).add_to(g_fue)
g_cal = folium.FeatureGroup(name='Mapa de calor', show=False)
sedes_js = []

# Diccionario ITEM -> entidad asignada (la más cercana, según df_det dedup).
# Usamos ITEM porque es único por fila (un mismo RUC puede tener múltiples
# registros en distintas fuentes: SUNAT, PREFA, COMPETENCIA).
_asig_dict = df_det.set_index(df_det[COL_DEMANDA['item']].astype(str))['ENTIDAD_ESTADO'].to_dict()

_otras_groups = []  # lista de grupos "otras en buffer" para controlar desde toggle global

for ent in entidades:
    col = color_map[ent]
    # Grupo principal: sedes + buffers + empresas ASIGNADAS (visible por defecto)
    grupo = folium.FeatureGroup(name='Entidad ' + ent, show=True)
    # Grupo secundario: empresas del buffer asignadas a OTRA entidad (contexto visual).
    # Nombre con prefijo especial "_hidden_" para poder OCULTARLO del panel de capas.
    grupo_otras = folium.FeatureGroup(name='_hidden_otras_' + ent, show=False, overlay=True, control=False)
    _otras_groups.append(grupo_otras)

    for _, sede in df_estado[df_estado[COL_ESTADO['cliente']] == ent].iterrows():
        nombre_sede = limpiar_sede(sede)
        distrito  = str(sede.get(COL_ESTADO['distrito'], ''))
        provincia = str(sede.get(COL_ESTADO['provincia'], ''))
        depto     = str(sede.get(COL_ESTADO['depto'], ''))

        res_sede = df_res[df_res['SEDE_ITEM'] == str(sede[COL_ESTADO['item']]).strip()]
        mst_km   = float(res_sede['METRAJE_MST_KM'].iloc[0]) if len(res_sede) > 0 else 0.0
        n_emp    = int(res_sede['TOTAL_EMPRESAS'].iloc[0])   if len(res_sede) > 0 else 0
        sede_item_str = str(sede[COL_ESTADO['item']]).strip()

        # VISUAL: todas las empresas del buffer (sin dedup), para pintar puntos en el mapa
        sub_visual = df_mapa_visual[df_mapa_visual['SEDE_ITEM'] == sede_item_str]
        empresas_coords = [[float(r[lat_d]), float(r[lon_d])] for _, r in sub_visual.iterrows()]

        # ASIGNADAS: solo las N empresas realmente asignadas a esta sede (df_det dedup)
        # Estas son las del popup (n_emp) y las únicas a las que se dibujan rutas reales.
        sub_asig = df_det[df_det['SEDE_ITEM'] == sede_item_str]
        empresas_asignadas_coords = [[float(r[lat_d]), float(r[lon_d])] for _, r in sub_asig.iterrows()]

        # Para conteo dinámico: por cada empresa en el buffer, qué entidad la tiene asignada
        _items_en_buffer = sub_visual[COL_DEMANDA['item']].astype(str).tolist()
        empresas_entidades = [_asig_dict.get(it, ent) for it in _items_en_buffer]

        sede_idx = len(sedes_js)
        sedes_js.append({
            'lat': float(sede[lat_e]), 'lon': float(sede[lon_e]),
            'nombre': nombre_sede, 'entidad': ent, 'color': col,
            'mst_km': mst_km, 'n_emp': n_emp,
            'empresas': empresas_asignadas_coords,   # para rutas reales (las 6 asignadas)
            'empresas_visuales': empresas_coords,    # para referencia (todas las del buffer)
            'emp_entidades': empresas_entidades,     # entidad asignada de cada empresa visual
        })

        folium.Circle(
            location=[sede[lat_e], sede[lon_e]], radius=BUFFER_METROS,
            color=col, fill=True, fill_opacity=0.08, weight=2, dash_array='8',
            tooltip='Buffer ' + str(BUFFER_METROS) + 'm - ' + ent + ' | ' + nombre_sede
        ).add_to(grupo)

        btn_label = 'Ver rutas reales por calle' if n_emp > 0 else 'Sin empresas en buffer'
        btn_style = ('background:' + col + ';color:white;border:none;padding:8px 0;'
                    'border-radius:6px;cursor:pointer;font-size:13px;font-weight:bold;width:100%')
        if n_emp == 0:
            btn_style = ('background:#ccc;color:#666;border:none;padding:8px 0;'
                        'border-radius:6px;font-size:13px;width:100%;cursor:default')

        popup_html = (
            "<div style='font-family:Arial;min-width:250px;max-width:300px'>"
            "<div style='background:" + col + ";color:white;padding:8px 12px;margin:-1px -1px 8px;border-radius:4px 4px 0 0'>"
            "<b style='font-size:14px'>" + ent + "</b></div>"
            "<b style='font-size:13px;color:#222'>" + nombre_sede + "</b><br><br>"
            "<span style='color:#555;font-size:12px'>" + str(sede[COL_ESTADO['direccion']]) + "</span><br>"
            "<span style='color:#777;font-size:11px'>" + distrito + " &bull; " + provincia + " &bull; " + depto + "</span><br><br>"
            "<div style='display:flex;gap:8px;margin-bottom:6px'>"
            "  <div style='flex:1;background:#f5f5f5;border-radius:6px;padding:6px;text-align:center'>"
            "    <div class='emp-count' data-sede='" + str(sede_idx) + "' style='font-size:18px;font-weight:bold;color:" + col + "'>" + str(n_emp) + "</div>"
            "    <div style='font-size:10px;color:#777'>empresas</div></div>"
            "  <div style='flex:1;background:#f5f5f5;border-radius:6px;padding:6px;text-align:center'>"
            "    <div style='font-size:18px;font-weight:bold;color:" + col + "'>" + str(mst_km) + "</div>"
            "    <div style='font-size:10px;color:#777'>km MST</div></div>"
            "</div>"
            "<button onclick='window.cargarRutas(" + str(sede_idx) + ")' style='" + btn_style + "'>"
            + btn_label +
            "</button>"
            "</div>"
        )

        folium.Marker(
            location=[sede[lat_e], sede[lon_e]],
            popup=folium.Popup(popup_html, max_width=320),
            tooltip=ent + ' - ' + nombre_sede,
            icon=folium.Icon(color='white', icon_color=col, icon='home', prefix='glyphicon')
        ).add_to(grupo)

        # Set de items asignados a ESTA sede (para diferenciar visualmente)
        _items_asignados = set(sub_asig[COL_DEMANDA['item']].astype(str))

        for _, emp in sub_visual.iterrows():
            fuente      = str(emp.get(COL_DEMANDA['fuente'], '')).upper()
            hex_col     = HEX_FUENTE.get(fuente, '#555')
            territorio  = str(emp.get(COL_DEMANDA['territorio'], ''))
            distrito_e  = str(emp.get(COL_DEMANDA['distrito'], ''))
            provincia_e = str(emp.get(COL_DEMANDA['provincia'], ''))
            dist_m      = int(emp.get('DISTANCIA_METROS', 0))
            # ¿Esta empresa está asignada a ESTA sede o a otra entidad por cercanía?
            _item_key = str(emp[COL_DEMANDA['item']])
            _es_asignada = _item_key in _items_asignados
            _ent_real    = _asig_dict.get(_item_key, ent)
            _badge_asig  = ("<span style='background:#2E7D32;color:white;padding:2px 8px;border-radius:4px;font-size:10px;font-weight:bold'>ASIGNADA</span>"
                            if _es_asignada
                            else "<span style='background:#777;color:white;padding:2px 8px;border-radius:4px;font-size:10px'>Asignada a " + _ent_real + "</span>")
            popup_emp = (
                "<div style='font-family:Arial;min-width:240px'>"
                "<b style='font-size:13px'>" + str(emp[COL_DEMANDA['cliente']]) + "</b><br>"
                "RUC: " + str(emp[COL_DEMANDA['ruc']]) + "<br>"
                "<span style='background:" + hex_col + ";color:white;padding:2px 8px;"
                "border-radius:4px;font-size:11px;font-weight:bold'>" + fuente + "</span> "
                + _badge_asig + "<br><br>"
                + str(emp.get(COL_DEMANDA['direccion'], '')) + "<br>"
                "<span style='color:#777;font-size:11px'>" + distrito_e + " &bull; " + provincia_e + "</span><br>"
                "Territorio: <b>" + territorio + "</b><br>"
                "Distancia a sede: <b>" + str(dist_m) + " m</b>"
                "</div>"
            )
            # Radio/opacidad distintos: asignadas sólidas, no-asignadas más pequeñas pero visibles
            _radio   = 8 if _es_asignada else 5
            _opacity = 1.0 if _es_asignada else 0.65
            _weight  = 1.5 if _es_asignada else 1.0
            _marker = folium.CircleMarker(
                location=[emp[lat_d], emp[lon_d]],
                radius=_radio, color='white', weight=_weight,
                fill=True, fill_color=hex_col, fill_opacity=_opacity,
                popup=folium.Popup(popup_emp, max_width=300),
                tooltip=fuente + ' - ' + str(emp[COL_DEMANDA['cliente']])[:35] + ' - ' + str(dist_m) + 'm'
            )
            # Asignadas al grupo principal (visible); otras al grupo secundario (oculto por defecto)
            if _es_asignada:
                _marker.add_to(grupo)
            else:
                _marker.add_to(grupo_otras)

    grupo.add_to(mapa)
    grupo_otras.add_to(mapa)

for _, emp in df_demanda_full[df_demanda_full['EN_BUFFER']=='FUERA'].iterrows():
    fuente  = str(emp.get(COL_DEMANDA['fuente'], '')).upper()
    hex_col = HEX_FUENTE.get(fuente, '#888')
    folium.CircleMarker(
        location=[emp[lat_d], emp[lon_d]], radius=4,
        color=hex_col, weight=0.8, fill=True, fill_color=hex_col, fill_opacity=0.4,
        tooltip='FUERA - ' + fuente + ' - ' + str(emp[COL_DEMANDA['cliente']])[:35]
    ).add_to(_cluster_fue)

HeatMap(list(zip(df_demanda[lat_d], df_demanda[lon_d])), radius=25, blur=20, min_opacity=0.4, max_zoom=14, gradient={'0.0':'rgba(0,0,255,0)','0.3':'#4575b4','0.5':'#fee090','0.7':'#f46d43','1.0':'#d73027'}).add_to(g_cal)
g_fue.add_to(mapa); g_cal.add_to(mapa)

mapa_id = mapa.get_name()
js_drag = [
    '<script>',
    'document.addEventListener("DOMContentLoaded", function() {',
    '  var st = document.createElement("style");',
    '  st.textContent = ".leaflet-popup-content-wrapper{max-width:340px!important;max-height:80vh!important;overflow-y:auto!important;} .leaflet-popup-content{width:auto!important;} .leaflet-popup-close-button{z-index:10!important;}";',
    '  document.head.appendChild(st);',
    '  function makeDraggable(popup) {',
    '    var el = popup.getElement(); if (!el || el._drag) return; el._drag = true;',
    '    var wrapper = el.querySelector(".leaflet-popup-content-wrapper");',
    '    var tip = el.querySelector(".leaflet-popup-tip-container");',
    '    var closeBtn = el.querySelector(".leaflet-popup-close-button");',
    '    if (!wrapper) return;',
    '    var dx=0, dy=0, mx=0, my=0, dragging=false;',
    '    wrapper.style.cursor = "move";',
    '    function aplicar(x, y) {',
    # movemos el wrapper Y el botón de cerrar juntos usando margin (no transform)
    '      wrapper.style.marginLeft = x + "px";',
    '      wrapper.style.marginTop = y + "px";',
    '      if (tip) { tip.style.marginLeft = x + "px"; tip.style.marginTop = y + "px"; }',
    '      if (closeBtn) { closeBtn.style.marginLeft = x + "px"; closeBtn.style.marginTop = y + "px"; }',
    '    }',
    '    wrapper.addEventListener("mousedown", function(e) {',
    '      if (e.target.closest(".leaflet-popup-close-button")) return;',
    '      e.preventDefault(); e.stopPropagation();',
    '      dragging=true; mx=e.clientX; my=e.clientY;',
    '      dx = parseInt(wrapper.style.marginLeft) || 0;',
    '      dy = parseInt(wrapper.style.marginTop) || 0;',
    '    });',
    '    document.addEventListener("mousemove", function(e) {',
    '      if (!dragging) return;',
    '      aplicar(dx + e.clientX - mx, dy + e.clientY - my);',
    '    });',
    '    document.addEventListener("mouseup", function() { dragging=false; });',
    '  }',
    '  var mapObj = window["'+mapa_id+'"];',
    '  if (mapObj) mapObj.on("popupopen", function(e) { setTimeout(function(){ makeDraggable(e.popup); }, 60); });',
    '});',
    '</script>'
]
mapa.get_root().html.add_child(folium.Element('\n'.join(js_drag)))

sedes_json_str = json.dumps(sedes_js, ensure_ascii=False)
js_parts = []
js_parts.append('<script>')
js_parts.append('var sedesData = ' + sedes_json_str + ';')
js_parts.append('window.cargarRutas = function(idx) {')
js_parts.append('  var sed = sedesData[idx];')
js_parts.append('  if (!sed || sed.empresas.length === 0) { alert("Sin empresas en buffer."); return; }')
js_parts.append('  var mapObj = window["' + mapa_id + '"];')
js_parts.append('  if (!mapObj) { alert("Mapa no encontrado"); return; }')
js_parts.append('  mapObj.eachLayer(function(l){ if(l.options && l.options._esRuta) mapObj.removeLayer(l); });')
js_parts.append('  var panel = document.getElementById("ruta-panel");')
js_parts.append('  var btnL  = document.getElementById("btn-limpiar");')
js_parts.append('  panel.style.display = "block"; if (btnL) btnL.style.display = "block";')
js_parts.append('  panel.style.borderColor = sed.color;')
js_parts.append('  panel.innerHTML = "<div id=\\"ruta-prog\\" style=\\"color:#666\\">Cargando <b><span id=\\"ruta-cnt\\">0</span>/"+sed.empresas.length+"</b> rutas...</div>";')
js_parts.append('  var cnt = 0;')
js_parts.append('  sed.empresas.forEach(function(emp) {')
js_parts.append('    var url = "https://router.project-osrm.org/route/v1/driving/" + sed.lon + "," + sed.lat + ";" + emp[1] + "," + emp[0] + "?overview=full&geometries=geojson";')
js_parts.append('    fetch(url).then(function(r){ return r.json(); }).then(function(data) {')
js_parts.append('      if (data.code === "Ok") {')
js_parts.append('        var pts = data.routes[0].geometry.coordinates.map(function(p){ return [p[1],p[0]]; });')
js_parts.append('        L.polyline(pts, {color:sed.color, weight:2.5, opacity:0.85, _esRuta:true}).addTo(mapObj);')
js_parts.append('      }')
js_parts.append('      cnt++; var el = document.getElementById("ruta-cnt"); if(el) el.textContent = cnt;')
js_parts.append('      if (cnt === sed.empresas.length) { panel.innerHTML = "<span style=color:#2E7D32;font-weight:bold>Rutas cargadas</span>"; setTimeout(function(){ panel.style.display="none"; }, 2000); }')
js_parts.append('    }).catch(function(){ cnt++; });')
js_parts.append('  });')
js_parts.append('};')
js_parts.append('window.limpiarRutas = function() {')
js_parts.append('  var mapObj = window["' + mapa_id + '"];')
js_parts.append('  if (mapObj) mapObj.eachLayer(function(l){ if(l.options && l.options._esRuta) mapObj.removeLayer(l); });')
js_parts.append('  var p = document.getElementById("ruta-panel"); if(p) p.style.display = "none";')
js_parts.append('  var b = document.getElementById("btn-limpiar"); if(b) b.style.display = "none";')
js_parts.append('};')
js_parts.append('</script>')
js_parts.append('<div id="ruta-panel" style="display:none;position:fixed;bottom:85px;right:10px;z-index:9999;background:white;padding:10px 14px;border-radius:10px;box-shadow:0 4px 16px rgba(0,0,0,0.18);font-family:Arial;font-size:12px;border-left:4px solid #1565C0;max-width:200px"></div>')
js_parts.append('<div id="btn-limpiar" onclick="window.limpiarRutas()" style="display:none;position:fixed;bottom:30px;right:10px;z-index:9999;background:#C62828;color:white;padding:12px 24px;border-radius:8px;box-shadow:0 3px 10px rgba(0,0,0,0.25);cursor:pointer;font-family:Arial;font-size:15px;font-weight:bold;">Limpiar rutas</div>')
mapa.get_root().html.add_child(folium.Element(''.join(js_parts)))

# --- JavaScript para conteo dinámico al filtrar entidades ---
# Al activar/desactivar capas de entidades, el número de empresas en cada popup
# se recalcula: cuenta empresas de esta sede + las que serían "reasignadas" si
# su entidad está oculta. Las capas cuyo nombre empieza con "Entidad " se
# consideran filtros de entidad.
js_dynamic = [
    '<script>',
    '(function() {',
    '  function iniciar() {',
    '    var mapObj = window["' + mapa_id + '"];',
    '    if (!mapObj) { setTimeout(iniciar, 200); return; }',  # reintentar si el mapa aún no existe
    '    var entidadesOcultas = {};',
    '    function recalcularConteos() {',
    '      var spans = document.querySelectorAll(".emp-count");',
    '      spans.forEach(function(sp) {',
    '        var idx = parseInt(sp.getAttribute("data-sede"));',
    '        if (isNaN(idx) || !sedesData[idx]) return;',
    '        var sed = sedesData[idx];',
    '        if (!sed.emp_entidades) return;',
    '        var count = 0;',
    '        sed.emp_entidades.forEach(function(entAsig) {',
    '          if (entAsig === sed.entidad) { count++; }',
    '          else if (entidadesOcultas[entAsig]) { count++; }',
    '        });',
    '        sp.textContent = count;',
    '      });',
    '    }',
    '    window._recalcConteos = recalcularConteos;',  # expuesto globalmente para debug
    '    mapObj.on("overlayadd", function(e) {',
    '      if (e.name && e.name.indexOf("Entidad ") === 0) {',
    '        delete entidadesOcultas[e.name.substring(8)];',
    '        recalcularConteos();',
    '      }',
    '    });',
    '    mapObj.on("overlayremove", function(e) {',
    '      if (e.name && e.name.indexOf("Entidad ") === 0) {',
    '        entidadesOcultas[e.name.substring(8)] = true;',
    '        recalcularConteos();',
    '      }',
    '    });',
    '    mapObj.on("popupopen", function() { setTimeout(recalcularConteos, 50); });',
    '    console.log("[conteo-dinamico] inicializado, mapa=" + "' + mapa_id + '");',
    '  }',
    '  if (document.readyState === "loading") {',
    '    document.addEventListener("DOMContentLoaded", iniciar);',
    '  } else { setTimeout(iniciar, 100); }',
    '})();',
    '</script>'
]
mapa.get_root().html.add_child(folium.Element('\n'.join(js_dynamic)))

# LEYENDA ORIGINAL
dentro_count = (df_demanda_full['EN_BUFFER'] == 'DENTRO').sum()
fuera_count  = (df_demanda_full['EN_BUFFER'] == 'FUERA').sum()
_ley = []
_ley.append("<div id='leyenda-panel' style='position:fixed;bottom:20px;left:10px;z-index:9999;background:white;padding:14px 16px;border-radius:12px;box-shadow:0 3px 14px rgba(0,0,0,0.18);font-family:Arial;font-size:12px;max-width:210px'>")
_ley.append("<div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:8px'>")
_ley.append("<b style='font-size:13px'>Leyenda</b>")
_ley.append("<span onclick='var p=document.getElementById(\"leyenda-body\");var b=document.getElementById(\"ley-btn\");if(p.style.display==\"none\"){p.style.display=\"block\";b.textContent=\"-\"}else{p.style.display=\"none\";b.textContent=\"+\"}' id='ley-btn' style='cursor:pointer;font-size:16px;font-weight:bold;color:#555;width:24px;height:24px;text-align:center;line-height:24px;background:#f0f0f0;border-radius:50%'>-</span>")
_ley.append("</div><div id='leyenda-body'>")
_ley.append("<b>Entidades (casa = sede):</b><br>")
for _e, _c in color_map.items():
    _ley.append("<span style='display:inline-block;width:11px;height:11px;border-radius:50%;background:" + _c + ";margin-right:6px;vertical-align:middle'></span><b>" + _e + "</b><br>")
_ley.append("<br><b>Demanda (punto = empresa):</b><br>")
_ley.append("<span style='background:#1565C0;color:white;padding:1px 7px;border-radius:3px;font-size:11px'>SUNAT</span>&nbsp;")
_ley.append("<span style='background:#2E7D32;color:white;padding:1px 7px;border-radius:3px;font-size:11px'>PREFACT.</span>&nbsp;")
_ley.append("<span style='background:#E65100;color:white;padding:1px 7px;border-radius:3px;font-size:11px'>COMPET.</span><br>")
_ley.append("<br><b>Dentro: " + str(dentro_count) + "</b> &nbsp; Fuera: " + str(fuera_count) + "<br>")
_ley.append("<br><i style='font-size:11px;color:#888'>Click sede para ver detalle<br>y cargar rutas reales</i>")
_ley.append("</div></div>")
mapa.get_root().html.add_child(folium.Element("".join(_ley)))

if nodos_analisis:
    COLOR_STATUS_N = {'PROPIO':'#FF6F00','ALQUILADO':'#6A1B9A','ALQUILADO - COUBICACION':'#0277BD'}
    COLOR_ESTADO_N = {'ON':'#2E7D32','WIN':'#1565C0'}
    grupos_nodos = {}
    for _est in sorted(set(n['estado'] for n in nodos_analisis)):
        grupos_nodos[_est] = folium.FeatureGroup(name='Nodos ['+_est+']', show=True)
    nodos_js_data = []
    for nodo in nodos_analisis:
        cn    = COLOR_STATUS_N.get(nodo['status'], '#37474F')
        c_est = COLOR_ESTADO_N.get(nodo['estado'], '#37474F')
        grupo = grupos_nodos.get(nodo['estado'], list(grupos_nodos.values())[0])
        ph = (
            "<div style='font-family:Arial;min-width:250px'>"
            "<div style='background:"+cn+";color:white;padding:8px 12px;border-radius:8px 8px 0 0;font-weight:bold'>Nodo: "+nodo['nombre']+"</div>"
            "<div style='padding:10px 12px;border:1px solid #eee;border-top:none;border-radius:0 0 8px 8px'>"
            "<div style='font-size:11px;color:#666;margin-bottom:8px'>"+nodo['ciudad']+" &bull; "+nodo['depto']+"<br>"+nodo['direccion']+"</div>"
            "<div style='display:flex;gap:6px;margin-bottom:8px'>"
            "<div style='flex:1;text-align:center;background:#f5f5f5;border-radius:6px;padding:6px'>"
            "<b style='font-size:20px;color:"+cn+"'>"+str(nodo['sedes_5km'])+"</b>"
            "<div style='font-size:10px;color:#888'>Sedes<br>Estado</div></div>"
            "<div style='flex:1;text-align:center;background:#E8F5E9;border-radius:6px;padding:6px'>"
            "<b style='font-size:20px;color:#2E7D32'>"+str(nodo['emp_ya_cubiertas'])+"</b>"
            "<div style='font-size:10px;color:#555'>Ya<br>cubiertas</div></div>"
            "<div style='flex:1;text-align:center;background:#FFF8E1;border-radius:6px;padding:6px'>"
            "<b style='font-size:20px;color:#F57F17'>"+str(nodo['emp_potenciales'])+"</b>"
            "<div style='font-size:10px;color:#555'>Otros clientes<br>potenciales</div></div>"
            "</div>"
            "<span style='background:"+cn+";color:white;padding:2px 8px;border-radius:3px;font-size:11px'>"+nodo['status']+"</span> "
            "<span style='background:"+c_est+";color:white;padding:2px 8px;border-radius:3px;font-size:11px'>"+nodo['estado']+"</span>"
            "</div></div>"
        )
        folium.Marker(
            location=[nodo['lat'], nodo['lon']], popup=folium.Popup(ph, max_width=260),
            tooltip='Nodo: '+nodo['nombre']+' | '+nodo['estado'],
            icon=folium.Icon(color='white', icon_color=c_est, icon='signal', prefix='glyphicon')
        ).add_to(grupo)
        nodos_js_data.append({'lat':nodo['lat'],'lon':nodo['lon'],'color':c_est,'r':NODOS_BUFFER_M})
    for g in grupos_nodos.values(): g.add_to(mapa)
    
    _njs = json.dumps(nodos_js_data, ensure_ascii=False)
    _js_buf = [
        '<script>(function(){',
        'var ND='+_njs+';var BC=[]; function clr(){BC.forEach(function(c){c.remove();});BC=[];}',
        'document.addEventListener("DOMContentLoaded",function(){',
        'var M=window["'+mapa_id+'"];if(!M)return;',
        'M.on("popupopen",function(e){clr();var ll=e.popup.getLatLng();if(!ll)return;',
        'ND.forEach(function(n){var d=M.distance(ll,L.latLng(n.lat,n.lon));',
        'if(d<10){var c=L.circle([n.lat,n.lon],{radius:n.r,color:n.color,weight:1.5,fill:true,fillColor:n.color,fillOpacity:0.08,dashArray:"6 4"}).addTo(M);BC.push(c);}});});',
        'M.on("popupclose",function(){clr();});',
        '});})()</script>',
    ]
    mapa.get_root().html.add_child(folium.Element(''.join(_js_buf)))

folium.LayerControl(collapsed=True).add_to(mapa)

# ==========================================
# MEJORAS UX: Dashboard de resumen visible
# ==========================================

# Datos agregados por entidad para el dashboard
_ent_stats = {}
for _e in entidades:
    _ses = df_res[df_res['ENTIDAD'] == _e]
    _ent_stats[_e] = {
        'n_sedes':  int(len(_ses)),
        'n_emp':    int(_ses['TOTAL_EMPRESAS'].sum()),
        'mst_km':   round(float(_ses['METRAJE_MST_KM'].sum()), 2),
        'color':    color_map[_e],
    }
_ent_stats_json = json.dumps(_ent_stats, ensure_ascii=False)

_ux_html = """
<style>
  #ux-dashboard {
    position: fixed; top: 60px; left: 60px; z-index: 9997;
    background: white; padding: 8px 10px; border-radius: 8px;
    box-shadow: 0 2px 12px rgba(0,0,0,0.15); font-family: Arial, sans-serif;
    font-size: 11px; max-width: 220px; min-width: 160px;
    max-height: calc(100vh - 140px); overflow-y: auto;
  }
  #ux-dashboard.collapsed { min-width: auto; }
  #ux-dashboard .dash-title { font-weight: bold; font-size: 11px; color: #333; display: flex; justify-content: space-between; align-items: center; gap: 10px; }
  #ux-dashboard:not(.collapsed) .dash-title { margin-bottom: 6px; }
  #ux-dashboard .dash-title span.dash-toggle { cursor: pointer; background: #f0f0f0; border-radius: 50%; width: 18px; height: 18px; text-align: center; line-height: 18px; font-size: 13px; font-weight: bold; flex-shrink: 0; }
  #ux-dashboard .dash-totals { display: flex; gap: 4px; margin-bottom: 6px; }
  #ux-dashboard .dash-card { flex: 1; background: #f8f9fa; border-radius: 5px; padding: 4px; text-align: center; }
  #ux-dashboard .dash-card .n { font-size: 13px; font-weight: bold; color: #1565C0; }
  #ux-dashboard .dash-card .l { font-size: 8px; color: #777; }
  #ux-dashboard .dash-ent-row { display: flex; align-items: center; gap: 5px; padding: 2px 0; font-size: 10px; border-bottom: 1px solid #f0f0f0; }
  #ux-dashboard .dash-ent-row .dot { width: 9px; height: 9px; border-radius: 50%; flex-shrink: 0; }
  #ux-dashboard .dash-ent-row .ent { font-weight: bold; flex: 1; }
  #ux-dashboard .dash-ent-row .stat { color: #555; font-size: 9px; }
  #ux-dashboard.collapsed .dash-body { display: none; }

  #ux-toggle-ctx {
    position: fixed; top: 12px; left: 60px; z-index: 9997;
    background: white; padding: 8px 14px; border-radius: 10px;
    box-shadow: 0 2px 12px rgba(0,0,0,0.15); font-family: Arial, sans-serif;
    font-size: 12px; cursor: pointer; display: flex; align-items: center; gap: 8px;
    border: 2px solid #ddd; transition: all 0.2s;
  }
  #ux-toggle-ctx:hover { border-color: #1565C0; }
  #ux-toggle-ctx.active { background: #1565C0; color: white; border-color: #0D47A1; }
  #ux-toggle-ctx .tc-dot { width: 10px; height: 10px; border-radius: 50%; background: #ccc; transition: background 0.2s; }
  #ux-toggle-ctx.active .tc-dot { background: #FFEB3B; }
</style>

<div id="ux-toggle-ctx" title="Al activar, se mostrarán también las empresas que están dentro del buffer de una sede pero que fueron asignadas a otra entidad más cercana (contexto de competencia).">
  <span class="tc-dot"></span>
  <span>Ver competidores en zona</span>
</div>

<div id="ux-dashboard" class="collapsed">
  <div class="dash-title">
    <span>📊 Resumen visible</span>
    <span class="dash-toggle" onclick="var d=document.getElementById('ux-dashboard');d.classList.toggle('collapsed');this.textContent=d.classList.contains('collapsed')?'+':'-';">+</span>
  </div>
  <div class="dash-body">
    <div class="dash-totals">
      <div class="dash-card"><div class="n" id="dash-t-ent">0</div><div class="l">Entidades</div></div>
      <div class="dash-card"><div class="n" id="dash-t-sed">0</div><div class="l">Sedes</div></div>
      <div class="dash-card"><div class="n" id="dash-t-emp">0</div><div class="l">Empresas</div></div>
    </div>
    <div id="dash-ent-list"></div>
  </div>
</div>

<script>
(function() {
  var entStats = __ENT_STATS__;

  function iniciarUX() {
    var mapObj = window["__MAPA_ID__"];
    if (!mapObj) { setTimeout(iniciarUX, 200); return; }

    var entidadesVisibles = {};
    Object.keys(entStats).forEach(function(e) { entidadesVisibles[e] = true; });

    function refrescarDashboard() {
      var tEnt=0, tSed=0, tEmp=0;
      var list = document.getElementById("dash-ent-list");
      list.innerHTML = "";
      Object.keys(entStats).sort().forEach(function(e) {
        var s = entStats[e];
        var vis = entidadesVisibles[e];
        if (vis) { tEnt++; tSed += s.n_sedes; tEmp += s.n_emp; }
        var row = document.createElement("div");
        row.className = "dash-ent-row";
        row.style.opacity = vis ? "1" : "0.35";
        row.innerHTML = '<div class="dot" style="background:' + s.color + '"></div>' +
                        '<div class="ent">' + e + '</div>' +
                        '<div class="stat">' + s.n_sedes + 's / ' + s.n_emp + 'e</div>';
        list.appendChild(row);
      });
      document.getElementById("dash-t-ent").textContent = tEnt;
      document.getElementById("dash-t-sed").textContent = tSed;
      document.getElementById("dash-t-emp").textContent = tEmp;
    }

    mapObj.on("overlayadd", function(e) {
      if (e.name && e.name.indexOf("Entidad ") === 0) {
        entidadesVisibles[e.name.substring(8)] = true;
        refrescarDashboard();
      }
    });
    mapObj.on("overlayremove", function(e) {
      if (e.name && e.name.indexOf("Entidad ") === 0) {
        entidadesVisibles[e.name.substring(8)] = false;
        refrescarDashboard();
      }
    });
    refrescarDashboard();

    // ===== TOGGLE DE CONTEXTO (mostrar/ocultar puntos "otras en buffer") =====
    // Los nombres de los FeatureGroups "otras" se inyectan desde Python.
    var btnCtx = document.getElementById("ux-toggle-ctx");
    var contextoVisible = false;
    var otrasLayerNames = __OTRAS_NAMES__;

    function obtenerOtrasLayers() {
      var layers = [];
      otrasLayerNames.forEach(function(name) {
        if (window[name]) layers.push(window[name]);
      });
      return layers;
    }

    btnCtx.addEventListener("click", function() {
      contextoVisible = !contextoVisible;
      btnCtx.classList.toggle("active", contextoVisible);
      btnCtx.querySelector("span:last-child").textContent =
        contextoVisible ? "Ocultar competidores" : "Ver competidores en zona";
      obtenerOtrasLayers().forEach(function(l) {
        if (contextoVisible) {
          if (!mapObj.hasLayer(l)) mapObj.addLayer(l);
        } else {
          if (mapObj.hasLayer(l)) mapObj.removeLayer(l);
        }
      });
    });
  }

  if (document.readyState === "loading") {
    document.addEventListener("DOMContentLoaded", iniciarUX);
  } else { setTimeout(iniciarUX, 100); }
})();
</script>
"""
_otras_js_names = json.dumps([g.get_name() for g in _otras_groups])
_ux_html = _ux_html.replace("__ENT_STATS__", _ent_stats_json)\
                   .replace("__MAPA_ID__", mapa_id)\
                   .replace("__OTRAS_NAMES__", _otras_js_names)
mapa.get_root().html.add_child(folium.Element(_ux_html))

out_mapa = os.path.join(SCRIPT_DIR, 'mapa_demanda.html')
mapa.save(out_mapa)
print(f"✅ Mapa guardado: {out_mapa}")

# ==========================================
# 8. EXPORTACIÓN A EXCEL
# ==========================================
print("\n📑 Generando reporte Excel con formato...")
out_excel = os.path.join(SCRIPT_DIR, 'resultados_demanda.xlsx')

# Join de nodos por coordenadas exactas (NO por RUC) para evitar asignar nodos
# de otra ciudad a empresas que comparten RUC con registros en otra región.
_COLS_NODO = ['NODO', 'NODO_ESTADO', 'NODO_STATUS']

def _build_nodo_lookup(df_nodos):
    if df_nodos is None or len(df_nodos) == 0:
        return None
    df_n = df_nodos.copy()
    df_n['_coord_key'] = df_n['LATITUD'].round(7).astype(str) + '|' + df_n['LONGITUD'].round(7).astype(str)
    return (
        df_n.sort_values('DIST_AL_NODO_M')
            .drop_duplicates(subset='_coord_key', keep='first')
            [['_coord_key', 'NODO', 'NODO_ESTADO', 'NODO_STATUS']]
            .set_index('_coord_key')
    )

def _merge_nodo(df, lat_col, lon_col, lookup):
    df = df.copy()
    if lookup is not None:
        df['_coord_key'] = df[lat_col].round(7).astype(str) + '|' + df[lon_col].round(7).astype(str)
        df = df.join(lookup, on='_coord_key')
        df.drop(columns='_coord_key', inplace=True)
    for c in _COLS_NODO:
        if c not in df.columns:
            df[c] = ''
        else:
            df[c] = df[c].fillna('')
    return df

nodo_lookup = _build_nodo_lookup(df_demanda_nodos)

cols_clave   = ['EN_BUFFER', 'ENTIDAD_ESTADO', 'SEDE_NOMBRE', 'SEDE_DIRECCION', 'DISTANCIA_METROS']
cols_empresa = [COL_DEMANDA['item'], COL_DEMANDA['ruc'], COL_DEMANDA['cliente'], COL_DEMANDA['direccion'],
                COL_DEMANDA['fuente'], COL_DEMANDA['territorio'], COL_DEMANDA['distrito'], COL_DEMANDA['provincia'], lat_d, lon_d]
cols_clave   = [c for c in cols_clave if c in df_demanda_full.columns]
cols_empresa = [c for c in cols_empresa if c in df_demanda_full.columns]
cols_resto   = [c for c in df_demanda_full.columns if c not in cols_clave + cols_empresa and not c.startswith('_')]

df_export_full = _merge_nodo(df_demanda_full, lat_d, lon_d, nodo_lookup)
cols_full = cols_clave + _COLS_NODO + [c for c in cols_empresa + cols_resto if c not in _COLS_NODO]
cols_full = [c for c in cols_full if c in df_export_full.columns]
df_export_full = df_export_full[cols_full].sort_values(['EN_BUFFER', 'ENTIDAD_ESTADO', 'SEDE_NOMBRE'], ascending=[False, True, True])

df_dentro_base = df_det[[c for c in (cols_clave + cols_empresa) if c in df_det.columns]].copy()
df_dentro_base = _merge_nodo(df_dentro_base, lat_d, lon_d, nodo_lookup)
cols_dentro = cols_clave + _COLS_NODO + [c for c in cols_empresa if c in df_dentro_base.columns and c not in _COLS_NODO]
cols_dentro = [c for c in cols_dentro if c in df_dentro_base.columns]
df_dentro = df_dentro_base[cols_dentro].sort_values(['ENTIDAD_ESTADO', 'SEDE_NOMBRE', 'DISTANCIA_METROS'], ascending=[True, True, True])

with pd.ExcelWriter(out_excel, engine='openpyxl') as w:
    df_ent.to_excel(w, sheet_name='Resumen por Entidad', index=False)
    df_res[[c for c in df_res.columns if not c.startswith('_')]].to_excel(w, sheet_name='Resumen por Sede', index=False)
    df_export_full.to_excel(w, sheet_name='Demanda Completa', index=False)
    df_dentro.to_excel(w, sheet_name='Solo Dentro Buffer', index=False)

wb = openpyxl.load_workbook(out_excel)
HDR = {'Resumen por Entidad':('1A237E','FFFFFF'), 'Resumen por Sede':('1A237E','FFFFFF'),
       'Demanda Completa':('1B5E20','FFFFFF'), 'Solo Dentro Buffer':('BF360C','FFFFFF')}
COLS_DESTACAR = {'EN_BUFFER', 'ENTIDAD_ESTADO', 'SEDE_NOMBRE', 'DISTANCIA_METROS'}
COLS_NODO_SET = set(_COLS_NODO)

for ws in wb.worksheets:
    bg, fg = HDR.get(ws.title, ('37474F', 'FFFFFF'))
    hdr_fill = PatternFill('solid', fgColor=bg)
    hdr_font = Font(bold=True, color=fg, size=10)
    header = [cell.value for cell in ws[1]]
    for cell in ws[1]:
        col_name = cell.value
        if col_name in COLS_NODO_SET:
            cell.fill = PatternFill('solid', fgColor='4A148C')
            cell.font = Font(bold=True, color='FFFFFF', size=10)
        elif col_name in COLS_DESTACAR:
            cell.fill = PatternFill('solid', fgColor='1565C0')
            cell.font = Font(bold=True, color='FFFFFF', size=10)
        else:
            cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    if ws.title in ('Demanda Completa', 'Solo Dentro Buffer'):
        if 'EN_BUFFER' in header:
            col_eb = header.index('EN_BUFFER') + 1
            for row in ws.iter_rows(min_row=2):
                val = row[col_eb - 1].value
                if val == 'DENTRO':
                    for cell in row: cell.fill = PatternFill('solid', fgColor='E8F5E9')
                elif val == 'FUERA':
                    for cell in row: cell.fill = PatternFill('solid', fgColor='F5F5F5')
    for col in ws.columns:
        w2 = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w2 + 4, 50)
    ws.freeze_panes = 'A2'

wb.save(out_excel)
print(f"✅ Excel guardado: {out_excel}")
print(f"🚀 Proceso completo. Archivos en: {SCRIPT_DIR}")